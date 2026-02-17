#!/usr/bin/env python3
"""
Directory Creator - Export Script
Generates CSV, Excel, PDF, and HTML files from directory data JSON.

Usage:
    python3 directory_export.py --input /tmp/directory_data.json
"""

import sys
import os
import json
import argparse
import csv
from datetime import datetime
import re

# Add virtual environment to Python path
VENV_SITE = os.path.expanduser("~/.claude/scripts/.venv/lib")
for d in os.listdir(VENV_SITE):
    sp = os.path.join(VENV_SITE, d, "site-packages")
    if os.path.isdir(sp) and sp not in sys.path:
        sys.path.insert(0, sp)

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Color scheme
TEAL = "#0D9488"
TEAL_DARK = "#0F766E"
TEAL_LIGHT = "#CCFBF1"

STATUS_COLORS = {
    "verified": {"bg": "#059669", "text": "#FFFFFF"},
    "needs_verification": {"bg": "#D97706", "text": "#FFFFFF"},
    "removed": {"bg": "#DC2626", "text": "#FFFFFF"}
}

DEFAULT_OUTPUT_DIR = "/mnt/c/Users/RonKlatt_3qsjg34/Desktop/Claude Code Plugin Output/Directory_Creator/"

CSV_COLUMNS = [
    "name", "slug", "category", "description", "address", "city", "state",
    "zip_code", "phone", "email", "website", "hours", "service_radius",
    "rating", "review_count", "services", "amenities", "service_areas",
    "facebook", "instagram", "linkedin", "twitter",
    "quality_score", "status", "source"
]


def latin_safe(text):
    if text is None:
        return ""
    return str(text).encode('latin-1', 'ignore').decode('latin-1')


def slugify(text):
    if not text:
        return "directory"
    text = text.lower()
    text = re.sub(r'[^a-z0-9]+', '_', text)
    return text.strip('_')


def score_color(score):
    if score >= 70:
        return "#059669"
    elif score >= 50:
        return "#0D9488"
    elif score >= 30:
        return "#D97706"
    else:
        return "#DC2626"


def flatten_business(biz):
    """Flatten a business dict for CSV export."""
    row = {}
    for col in CSV_COLUMNS:
        if col in ("facebook", "instagram", "linkedin", "twitter"):
            social = biz.get("social_links", {}) or {}
            row[col] = social.get(col, "")
        elif col in ("services", "amenities", "service_areas"):
            val = biz.get(col, [])
            row[col] = "; ".join(val) if isinstance(val, list) else str(val or "")
        else:
            row[col] = biz.get(col, "")
    return row


def generate_csv(data, output_path):
    """Generate DirectoryGenius-compatible CSV."""
    businesses = [b for b in data.get("businesses", []) if b.get("status") != "removed"]

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for biz in businesses:
            writer.writerow(flatten_business(biz))

    print(f"  CSV: {len(businesses)} listings exported")


def generate_excel(data, output_path):
    """Generate multi-sheet Excel workbook."""
    wb = Workbook()
    businesses = data.get("businesses", [])
    active_businesses = [b for b in businesses if b.get("status") != "removed"]
    stats = data.get("statistics", {})
    directory_name = data.get("directory_name", "Business Directory")

    # --- Sheet 1: Listings ---
    ws = wb.active
    ws.title = "Listings"

    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    headers = ["#", "Name", "Category", "City", "State", "Phone", "Website",
               "Rating", "Reviews", "Score", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, biz in enumerate(active_businesses, 2):
        values = [
            row_idx - 1,
            biz.get("name", ""),
            biz.get("category", ""),
            biz.get("city", ""),
            biz.get("state", ""),
            biz.get("phone", ""),
            biz.get("website", ""),
            biz.get("rating", ""),
            biz.get("review_count", ""),
            biz.get("quality_score", 0),
            biz.get("status", "")
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center')

    col_widths = [5, 35, 20, 15, 8, 18, 35, 8, 10, 8, 18]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # --- Sheet 2: Statistics ---
    ws2 = wb.create_sheet("Statistics")
    ws2.cell(row=1, column=1, value=directory_name).font = Font(bold=True, size=14, color="0F766E")
    ws2.cell(row=2, column=1, value=f"Generated: {data.get('generated_date', '')}")
    ws2.cell(row=3, column=1, value=f"Niche: {data.get('niche_type', '')}")
    ws2.cell(row=4, column=1, value=f"Geography: {data.get('geography', '')}")
    ws2.cell(row=5, column=1, value=f"Depth: {data.get('depth', '')}")

    ws2.cell(row=7, column=1, value="Statistics").font = Font(bold=True, size=12, color="0F766E")
    stat_rows = [
        ("Total Found", stats.get("total_found", 0)),
        ("Verified", stats.get("verified", 0)),
        ("Needs Verification", stats.get("needs_verification", 0)),
        ("Duplicates Removed", stats.get("duplicates_removed", 0)),
        ("Low Quality Removed", stats.get("low_quality_removed", 0)),
    ]
    for idx, (label, value) in enumerate(stat_rows, 8):
        ws2.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=idx, column=2, value=value)

    methodology = data.get("methodology", {})
    if methodology:
        ws2.cell(row=14, column=1, value="Methodology").font = Font(bold=True, size=12, color="0F766E")
        meth_rows = [
            ("Agents Dispatched", methodology.get("agents_dispatched", 0)),
            ("Total Searches", methodology.get("total_searches", 0)),
            ("Websites Crawled", methodology.get("websites_crawled", 0)),
            ("Duration (minutes)", methodology.get("duration_minutes", 0)),
        ]
        for idx, (label, value) in enumerate(meth_rows, 15):
            ws2.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws2.cell(row=idx, column=2, value=value)

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15

    # --- Sheet 3: Categories ---
    ws3 = wb.create_sheet("Categories")
    categories = {}
    for biz in active_businesses:
        cat = biz.get("category", "Uncategorized") or "Uncategorized"
        if cat not in categories:
            categories[cat] = {"count": 0, "avg_score": 0, "scores": []}
        categories[cat]["count"] += 1
        categories[cat]["scores"].append(biz.get("quality_score", 0))

    for cat in categories:
        scores = categories[cat]["scores"]
        categories[cat]["avg_score"] = round(sum(scores) / len(scores), 1) if scores else 0

    cat_headers = ["Category", "Count", "Avg Score"]
    for col_idx, header in enumerate(cat_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, (cat, info) in enumerate(sorted(categories.items(), key=lambda x: -x[1]["count"]), 2):
        ws3.cell(row=row_idx, column=1, value=cat)
        ws3.cell(row=row_idx, column=2, value=info["count"])
        ws3.cell(row=row_idx, column=3, value=info["avg_score"])

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 12

    # --- Sheet 4: Geography ---
    ws4 = wb.create_sheet("Geography")
    cities = {}
    for biz in active_businesses:
        city = biz.get("city", "Unknown") or "Unknown"
        if city not in cities:
            cities[city] = {"count": 0, "avg_score": 0, "scores": []}
        cities[city]["count"] += 1
        cities[city]["scores"].append(biz.get("quality_score", 0))

    for city in cities:
        scores = cities[city]["scores"]
        cities[city]["avg_score"] = round(sum(scores) / len(scores), 1) if scores else 0

    geo_headers = ["City", "State", "Count", "Avg Score"]
    for col_idx, header in enumerate(geo_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    for city, info in sorted(cities.items(), key=lambda x: -x[1]["count"]):
        ws4.cell(row=row_idx, column=1, value=city)
        state = ""
        for biz in active_businesses:
            if biz.get("city") == city:
                state = biz.get("state", "")
                break
        ws4.cell(row=row_idx, column=2, value=state)
        ws4.cell(row=row_idx, column=3, value=info["count"])
        ws4.cell(row=row_idx, column=4, value=info["avg_score"])
        row_idx += 1

    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 8
    ws4.column_dimensions['C'].width = 10
    ws4.column_dimensions['D'].width = 12

    wb.save(output_path)
    print(f"  Excel: 4 sheets ({len(active_businesses)} listings)")


def generate_pdf(data, output_path):
    """Generate PDF summary report."""
    directory_name = data.get("directory_name", "Business Directory")
    businesses = [b for b in data.get("businesses", []) if b.get("status") != "removed"]
    stats = data.get("statistics", {})

    class DirectoryPDF(FPDF):
        def header(self):
            self.set_fill_color(13, 148, 136)
            self.rect(0, 0, 210, 40, 'F')
            self.set_text_color(255, 255, 255)
            self.set_font('Helvetica', 'B', 20)
            self.cell(0, 15, '', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.cell(0, 10, latin_safe(directory_name), new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            self.set_font('Helvetica', '', 11)
            subtitle = f"{data.get('niche_type', '')} | {data.get('geography', '')} | {data.get('generated_date', '')}"
            self.cell(0, 8, latin_safe(subtitle), new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            self.ln(10)

        def footer(self):
            self.set_y(-15)
            self.set_font('Helvetica', 'I', 8)
            self.set_text_color(128, 128, 128)
            self.cell(0, 10, f'Page {self.page_no()}', new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')

    pdf = DirectoryPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Statistics summary
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(15, 118, 110)
    pdf.cell(0, 10, 'Directory Statistics', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    pdf.set_font('Helvetica', '', 11)
    pdf.set_text_color(0, 0, 0)

    stat_items = [
        ("Total Found", stats.get("total_found", 0)),
        ("Verified", stats.get("verified", 0)),
        ("Needs Verification", stats.get("needs_verification", 0)),
        ("Duplicates Removed", stats.get("duplicates_removed", 0)),
        ("Low Quality Removed", stats.get("low_quality_removed", 0)),
    ]
    for label, val in stat_items:
        pdf.cell(80, 8, latin_safe(f"{label}: {val}"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    methodology = data.get("methodology", {})
    if methodology:
        pdf.ln(3)
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 8, 'Methodology', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(90, 7, latin_safe(f"Agents: {methodology.get('agents_dispatched', 0)}"), new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.cell(90, 7, latin_safe(f"Searches: {methodology.get('total_searches', 0)}"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(90, 7, latin_safe(f"Websites crawled: {methodology.get('websites_crawled', 0)}"), new_x=XPos.RIGHT, new_y=YPos.TOP)
        pdf.cell(90, 7, latin_safe(f"Duration: {methodology.get('duration_minutes', 0)} min"), new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.ln(8)

    # Top listings table
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_text_color(15, 118, 110)
    pdf.cell(0, 10, 'Top Listings', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    # Table header
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_fill_color(13, 148, 136)
    pdf.set_text_color(255, 255, 255)
    col_widths = [8, 50, 25, 30, 18, 15, 14, 15, 15]
    col_headers = ["#", "Name", "Category", "City", "Phone", "Rating", "Rev", "Score", "Status"]
    for w, h in zip(col_widths, col_headers):
        pdf.cell(w, 7, h, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
    pdf.ln()

    pdf.set_font('Helvetica', '', 7)
    pdf.set_text_color(0, 0, 0)

    top = sorted(businesses, key=lambda x: x.get("quality_score", 0), reverse=True)[:50]
    for idx, biz in enumerate(top, 1):
        vals = [
            str(idx),
            latin_safe(biz.get("name", "")[:28]),
            latin_safe(biz.get("category", "")[:14]),
            latin_safe(biz.get("city", "")[:16]),
            latin_safe(str(biz.get("phone", ""))[-10:] if biz.get("phone") else ""),
            str(biz.get("rating", "") or ""),
            str(biz.get("review_count", "") or ""),
            str(biz.get("quality_score", 0)),
            biz.get("status", "")[:8]
        ]
        for w, v in zip(col_widths, vals):
            pdf.cell(w, 6, v, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C')
        pdf.ln()

        if pdf.get_y() > 270:
            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_fill_color(13, 148, 136)
            pdf.set_text_color(255, 255, 255)
            for w, h in zip(col_widths, col_headers):
                pdf.cell(w, 7, h, border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='C', fill=True)
            pdf.ln()
            pdf.set_font('Helvetica', '', 7)
            pdf.set_text_color(0, 0, 0)

    pdf.output(output_path)
    print(f"  PDF: {len(top)} top listings")


def generate_html(data, output_path):
    """Generate styled HTML directory preview with filterable table."""
    directory_name = data.get("directory_name", "Business Directory")
    businesses = [b for b in data.get("businesses", []) if b.get("status") != "removed"]
    stats = data.get("statistics", {})
    niche = data.get("niche_type", "")
    geography = data.get("geography", "")
    generated_date = data.get("generated_date", datetime.now().strftime("%Y-%m-%d"))
    depth = data.get("depth", "standard").upper()

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{directory_name}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            line-height: 1.6; color: #1f2937;
            background: linear-gradient(135deg, #f0fdfa 0%, #fff 100%);
            padding: 2rem;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; background: white;
            border-radius: 16px; box-shadow: 0 20px 60px rgba(0,0,0,0.1); overflow: hidden; }}
        .header {{
            background: linear-gradient(135deg, {TEAL} 0%, {TEAL_DARK} 100%);
            color: white; padding: 3rem 2rem; text-align: center;
        }}
        .header h1 {{ font-size: 2.5rem; margin-bottom: 0.5rem; }}
        .header .subtitle {{ font-size: 1.2rem; opacity: 0.95; margin-bottom: 1.5rem; }}
        .meta {{ display: flex; justify-content: center; gap: 2rem; flex-wrap: wrap; }}
        .meta-item {{ background: rgba(255,255,255,0.15); padding: 0.5rem 1rem;
            border-radius: 8px; }}
        .meta-label {{ opacity: 0.8; font-size: 0.85rem; }}
        .meta-value {{ font-weight: 600; font-size: 1.1rem; }}
        .content {{ padding: 2rem; }}
        .section {{ margin-bottom: 2rem; }}
        .section-title {{ font-size: 1.5rem; color: {TEAL_DARK}; margin-bottom: 1rem;
            padding-bottom: 0.5rem; border-bottom: 3px solid {TEAL}; }}
        .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 1rem; margin-bottom: 2rem; }}
        .stat-card {{ background: #f9fafb; padding: 1.5rem; border-radius: 8px; text-align: center; }}
        .stat-value {{ font-size: 2rem; font-weight: 700; color: {TEAL}; }}
        .stat-label {{ font-size: 0.85rem; color: #6b7280; }}
        .filter-bar {{ display: flex; gap: 1rem; margin-bottom: 1.5rem; flex-wrap: wrap; }}
        .filter-bar input, .filter-bar select {{ padding: 0.5rem 1rem; border: 1px solid #d1d5db;
            border-radius: 8px; font-size: 0.95rem; }}
        .filter-bar input {{ flex: 1; min-width: 200px; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th {{ background: {TEAL}; color: white; padding: 0.8rem; text-align: left;
            font-size: 0.85rem; text-transform: uppercase; cursor: pointer; white-space: nowrap; }}
        th:hover {{ background: {TEAL_DARK}; }}
        td {{ padding: 0.8rem; border-bottom: 1px solid #e5e7eb; font-size: 0.9rem; }}
        tr:hover {{ background: #f0fdfa; }}
        .status {{ display: inline-block; padding: 0.2rem 0.6rem; border-radius: 4px;
            font-size: 0.8rem; font-weight: 600; text-transform: uppercase; }}
        .status-verified {{ background: #D1FAE5; color: #065F46; }}
        .status-needs_verification {{ background: #FEF3C7; color: #92400E; }}
        .score-bar {{ width: 60px; height: 6px; background: #e5e7eb; border-radius: 3px;
            display: inline-block; margin-left: 0.5rem; vertical-align: middle; }}
        .score-fill {{ height: 100%; border-radius: 3px; }}
        a {{ color: {TEAL}; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        .footer {{ text-align: center; padding: 2rem; color: #6b7280; font-size: 0.9rem;
            border-top: 1px solid #e5e7eb; }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <h1>{directory_name}</h1>
        <div class="subtitle">{niche} | {geography}</div>
        <div class="meta">
            <div class="meta-item"><div class="meta-label">Generated</div><div class="meta-value">{generated_date}</div></div>
            <div class="meta-item"><div class="meta-label">Depth</div><div class="meta-value">{depth}</div></div>
            <div class="meta-item"><div class="meta-label">Listings</div><div class="meta-value">{len(businesses)}</div></div>
            <div class="meta-item"><div class="meta-label">Verified</div><div class="meta-value">{stats.get('verified', 0)}</div></div>
        </div>
    </div>
    <div class="content">
        <div class="section">
            <div class="stats-grid">
                <div class="stat-card"><div class="stat-value">{stats.get('total_found', 0)}</div><div class="stat-label">Total Found</div></div>
                <div class="stat-card"><div class="stat-value">{stats.get('verified', 0)}</div><div class="stat-label">Verified</div></div>
                <div class="stat-card"><div class="stat-value">{stats.get('needs_verification', 0)}</div><div class="stat-label">Needs Review</div></div>
                <div class="stat-card"><div class="stat-value">{stats.get('duplicates_removed', 0)}</div><div class="stat-label">Dupes Removed</div></div>
            </div>
        </div>
        <div class="section">
            <h2 class="section-title">Directory Listings</h2>
            <div class="filter-bar">
                <input type="text" id="search" placeholder="Search by name, city, or category..." onkeyup="filterTable()">
                <select id="statusFilter" onchange="filterTable()">
                    <option value="">All Statuses</option>
                    <option value="verified">Verified</option>
                    <option value="needs_verification">Needs Verification</option>
                </select>
                <select id="categoryFilter" onchange="filterTable()">
                    <option value="">All Categories</option>
"""

    categories = sorted(set(b.get("category", "Uncategorized") or "Uncategorized" for b in businesses))
    for cat in categories:
        html += f'                    <option value="{cat}">{cat}</option>\n'

    html += """                </select>
            </div>
            <table id="directoryTable">
                <thead>
                    <tr>
                        <th onclick="sortTable(0)">#</th>
                        <th onclick="sortTable(1)">Name</th>
                        <th onclick="sortTable(2)">Category</th>
                        <th onclick="sortTable(3)">City</th>
                        <th onclick="sortTable(4)">Phone</th>
                        <th onclick="sortTable(5)">Rating</th>
                        <th onclick="sortTable(6)">Reviews</th>
                        <th onclick="sortTable(7)">Score</th>
                        <th onclick="sortTable(8)">Status</th>
                    </tr>
                </thead>
                <tbody>
"""

    sorted_biz = sorted(businesses, key=lambda x: x.get("quality_score", 0), reverse=True)
    for idx, biz in enumerate(sorted_biz, 1):
        name = biz.get("name", "")
        website = biz.get("website", "")
        name_cell = f'<a href="{website}" target="_blank">{name}</a>' if website else name
        category = biz.get("category", "") or ""
        city = biz.get("city", "")
        phone = biz.get("phone", "")
        rating = biz.get("rating", "")
        reviews = biz.get("review_count", "")
        score = biz.get("quality_score", 0)
        status = biz.get("status", "")
        status_class = f"status-{status}" if status else ""
        sc = score_color(score)

        html += f"""                    <tr data-category="{category}" data-status="{status}">
                        <td>{idx}</td>
                        <td>{name_cell}</td>
                        <td>{category}</td>
                        <td>{city}</td>
                        <td>{phone}</td>
                        <td>{rating or '-'}</td>
                        <td>{reviews or '-'}</td>
                        <td>{score} <span class="score-bar"><span class="score-fill" style="width:{score}%;background:{sc};"></span></span></td>
                        <td><span class="status {status_class}">{status.replace('_', ' ')}</span></td>
                    </tr>
"""

    html += f"""                </tbody>
            </table>
        </div>
    </div>
    <div class="footer">
        Generated by Directory Creator on {generated_date}<br>
        ROK Plugin Marketplace - Directory Creator v1.0
    </div>
</div>
<script>
function filterTable() {{
    const search = document.getElementById('search').value.toLowerCase();
    const statusFilter = document.getElementById('statusFilter').value;
    const categoryFilter = document.getElementById('categoryFilter').value;
    const rows = document.querySelectorAll('#directoryTable tbody tr');
    rows.forEach(row => {{
        const text = row.textContent.toLowerCase();
        const status = row.dataset.status;
        const category = row.dataset.category;
        const matchSearch = !search || text.includes(search);
        const matchStatus = !statusFilter || status === statusFilter;
        const matchCategory = !categoryFilter || category === categoryFilter;
        row.style.display = matchSearch && matchStatus && matchCategory ? '' : 'none';
    }});
}}
let sortDir = {{}};
function sortTable(colIdx) {{
    const table = document.getElementById('directoryTable');
    const tbody = table.querySelector('tbody');
    const rows = Array.from(tbody.querySelectorAll('tr'));
    sortDir[colIdx] = !sortDir[colIdx];
    rows.sort((a, b) => {{
        let aVal = a.cells[colIdx].textContent.trim();
        let bVal = b.cells[colIdx].textContent.trim();
        const aNum = parseFloat(aVal);
        const bNum = parseFloat(bVal);
        if (!isNaN(aNum) && !isNaN(bNum)) {{
            return sortDir[colIdx] ? aNum - bNum : bNum - aNum;
        }}
        return sortDir[colIdx] ? aVal.localeCompare(bVal) : bVal.localeCompare(aVal);
    }});
    rows.forEach(row => tbody.appendChild(row));
}}
</script>
</body>
</html>"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"  HTML: {len(businesses)} listings with filtering")


def main():
    parser = argparse.ArgumentParser(description='Export directory data to CSV, Excel, PDF, and HTML')
    parser.add_argument('--input', required=True, help='Path to directory_data.json')
    parser.add_argument('--output-dir', default=DEFAULT_OUTPUT_DIR, help='Output directory')

    args = parser.parse_args()

    try:
        with open(args.input, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON: {e}", file=sys.stderr)
        sys.exit(1)

    if data.get('type') != 'directory_data':
        print(f"Warning: Expected type 'directory_data', got '{data.get('type')}'", file=sys.stderr)

    os.makedirs(args.output_dir, exist_ok=True)

    name_slug = slugify(data.get("directory_name", "directory"))
    date_str = data.get("generated_date", datetime.now().strftime("%Y-%m-%d"))
    base = f"directory_{name_slug}_{date_str}"

    try:
        csv_path = os.path.join(args.output_dir, f"{base}.csv")
        print(f"Generating CSV: {csv_path}")
        generate_csv(data, csv_path)

        xlsx_path = os.path.join(args.output_dir, f"{base}.xlsx")
        print(f"Generating Excel: {xlsx_path}")
        generate_excel(data, xlsx_path)

        pdf_path = os.path.join(args.output_dir, f"{base}.pdf")
        print(f"Generating PDF: {pdf_path}")
        generate_pdf(data, pdf_path)

        html_path = os.path.join(args.output_dir, f"{base}.html")
        print(f"Generating HTML: {html_path}")
        generate_html(data, html_path)

        print(f"\nExport complete!")
        print(f"\nOutput files:")
        print(f"  - {csv_path}")
        print(f"  - {xlsx_path}")
        print(f"  - {pdf_path}")
        print(f"  - {html_path}")

    except Exception as e:
        print(f"Error during export: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
